"""
Python port of the MPNDailyAutomate VBA macro.

Reads the e-Builder "Systemwide Daily Material Placement Summary" export
(.xls/.xlsm in the same layout as MacroTest.xlsm), extracts the relevant
columns for each placement row, computes Contract Number and Route,
sorts by Contract then Date, and writes the result to a separate Excel file.

Usage:
    python extract_mpn.py <input_file> <output_file>
"""
import os
import smtplib
import sys
import re
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import json
from openpyxl.styles import PatternFill

with open("config.json", "r") as f:
    config = json.load(f)

EMAIL_ADDRESS = config["email_address"]
EMAIL_PASSWORD = config["email_password"]

DEST_PATH = Path(config["destination_path"])
TRACKER_PATH = Path(config["tracker_path"])

# EMAIL_ADDRESS = os.environ["EMAIL_ADDRESS"]
# EMAIL_PASSWORD = os.environ["EMAIL_PASSWORD"]
# EMAIL_ADDRESS = "rowanawhite@icloud.com"
# EMAIL_PASSWORD = "dbjp-rgvh-yqga-wswx"

BASE_DIR = Path(__file__).parent

# DEST_PATH = BASE_DIR / "TestDest2.xlsx"
# TRACKER_PATH = BASE_DIR / "TestTracker2.xlsx"
# DEST_PATH = Path(r"G:\Shared drives\MEC Shared Drive\08 MITScan\MITScan\Deficiencies_(Rowan's)\TestDest2.xlsx")
# TRACKER_PATH = Path(r"G:\Shared drives\MEC Shared Drive\08 MITScan\MITScan\Deficiencies_(Rowan's)\TestTracker2.xlsx")

INPUT_PATH = BASE_DIR / "MacroTest.xlsm"
# -----------------------------------------------------------------------
# Route lookup, mirrors the If/ElseIf chain in the VBA macro.
# Add new contract numbers -> route names here.
# -----------------------------------------------------------------------
ROUTE_LOOKUP = {
    4673: "I-490", 4743: "I-490", 4714: "I-490", 4738: "I-490",
    4746: "I-490", 4727: "I-490", 4758: "I-490", 4622: "I-490",
    4827: "I-94", 4914: "I-94",
    4464: "I-57", 4475: "I-57",
    4722: "EOWA",
    4587: "M5",
    4937: "Cal Sag",
    4940: "SW",
    4904: "I-88", 4909: "I-88", 4917: "I-88 Dekalb", 4938: "I-88", 4939: "I-88",
    4915: "I-90 Genoa",
    4916: " I-355 Roosevelt",
    4918: "Plaza ORD",
    4804: "Plaza 47",
    4872: "I-355",
    4931: "Systemwide Patching",
    4732: " Railroad", 4736: " Railroad", 4737: " Railroad",
    4866: "Plaza Riverside",
}

# I-294 contract numbers (large OR chain in the VBA)
_I294_CONTRACTS = [
    4582, 4339, 4428, 4431, 4458, 4491, 4496, 4342, 4519, 4819, 4518, 4598,
    4814, 4533, 4517, 4831, 4832, 4833, 4834, 4835, 4836, 4884, 4594, 4801,
    4856, 4860, 4869,
]
for c in _I294_CONTRACTS:
    ROUTE_LOOKUP[c] = "I-294"

# Unit abbreviation map, mirrors the VBA InStr() chain
UNIT_ABBREVIATIONS = {
    "cubic yards": "CY",
    "pounds": "lbs",
    "square yards": "SY",
    "lineal feet": "LF",
    "square feet": "SF",
}


def abbreviate_unit(unit_text):
    """Return abbreviation for a unit string, or the original text if unknown."""
    if not unit_text:
        return unit_text
    for key, abv in UNIT_ABBREVIATIONS.items():
        if key in unit_text:
            return abv
    return unit_text


def extract_contract_number(project_name_text):
    """
    Pull the contract number out of a 'Project Name: 4975C ...' string.

    The contract number is the leading run of digits immediately following
    'Project Name:' (e.g. '4975C I-90 and I-355 ...' -> 4975).
    """
    if not project_name_text:
        return None
    match = re.search(r"Project Name:\s*(\d+)", project_name_text)
    if not match:
        return None
    return int(match.group(1))


def find_header_row(ws):
    """Locate the row containing the column headers (the row with 'Process Counter')."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and "Process Counter" in val:
                return row
    raise ValueError("Could not find header row (no 'Process Counter' cell found).")


def map_columns(ws, header_row):
    """Map the needed source columns by header text, like the VBA InStr loop."""
    wanted = {
        "Date of Placement (Proposed)": "date",
        "Description of Work": "activity",
        "Mix Design": "material",
        "Quantity Placed (Proposed)": "quantity",
        "Unit of Measurement (Proposed)": "units",
        "Material Producer": "source",
        "Company Placing the Material": "contractor",
        "Time of Placement (Proposed)": "time",
        "Location Description": "location",
        "Process Counter": "process_counter",
    }

    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if not isinstance(val, str):
            continue
        for header_text, key in wanted.items():
            if header_text in val and key not in col_map:
                col_map[key] = col

    if "time" not in col_map and "date" in col_map:
        col_map["time"] = col_map["date"] + 1

    return col_map


def extract_records(ws, header_row, col_map):
    """
    Walk rows below the header row.

    Mirrors the VBA: a row starting with 'Project Name:' (in column A or B)
    sets the active Contract Number / Route for following data rows.
    Data rows are identified by having a value in column C (Process Counter).
    """
    records = []
    current_contract = None
    current_route = None

    for row in range(header_row + 1, ws.max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value
        text = ""
        if isinstance(a_val, str):
            text = a_val
        elif isinstance(b_val, str):
            text = b_val

        if "Project Name" in text:
            contract_num = extract_contract_number(text)
            current_contract = contract_num
            current_route = ROUTE_LOOKUP.get(contract_num, "") if contract_num else ""
            continue

        # Data rows have a Process Counter value in column C
        c_val = ws.cell(row=row, column=3).value
        if c_val in (None, ""):
            continue

        def get(key):
            col = col_map.get(key)
            return ws.cell(row=row, column=col).value if col else None

        record = {
            "Date": get("date"),
            "Contract": current_contract,
            "Route": current_route,
            "Activity": get("activity"),
            "Material": get("material"),
            "Quantity": get("quantity"),
            "Units": abbreviate_unit(get("units")),
            "Source": get("source"),
            "Contractor": get("contractor"),
            "Time": get("time"),
            "Location": get("location"),
            "Process Counter": get("process_counter"),
        }
        records.append(record)

    return records


def parse_date_for_sort(value):
    """Best-effort parse of date values (string 'MM.DD.YYYY' or datetime) for sorting."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%m.%d.%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return datetime.min


def sort_records(records):
    """Sort by Contract Number, then by Date - matches the VBA two-pass sort."""
    return sorted(
        records,
        key=lambda r: (
            r["Contract"] if r["Contract"] is not None else 0,
            parse_date_for_sort(r["Date"]),
        ),
    )


OUTPUT_COLUMNS = [
    "Date",
    "Contract",
    "Route",
    "Activity",
    "Material",
    "Quantity",
    "Units",
    "Source",
    "Contractor",
    "Time",
    "Location",
    "Process Counter",
]


def write_output(records, output_path, source_ws=None, append_to_source=False):
    """
    Write extracted/sorted records.

    If append_to_source is True and source_ws is provided, the extracted
    columns are pasted to the side of the original sheet starting at
    column AE (matching the VBA layout), and the workbook is saved to
    output_path. Otherwise, a fresh workbook with just the extracted
    table is created.
    """
    if append_to_source and source_ws is not None:
        wb = source_ws.parent
        ws = source_ws
        start_col = 31  # column AE
        header_row = find_header_row(ws)

        for j, col_name in enumerate(OUTPUT_COLUMNS):
            ws.cell(row=header_row, column=start_col + j, value=col_name)

        for i, rec in enumerate(records):
            for j, col_name in enumerate(OUTPUT_COLUMNS):
                ws.cell(row=header_row + 1 + i, column=start_col + j, value=rec[col_name])

        wb.save(output_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted"

        for j, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=1, column=j, value=col_name)

        for i, rec in enumerate(records, start=2):
            for j, col_name in enumerate(OUTPUT_COLUMNS, start=1):
                ws.cell(row=i, column=j, value=rec[col_name])

        # Auto-size columns a bit
        for j, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(j)].width = max(12, len(col_name) + 2)

        wb.save(output_path)


def parse_time_value(value):
    """Convert a time string like '7:00 AM' to a datetime.time, matching destination format."""
    if isinstance(value, str):
        for fmt in ("%I:%M %p", "%H:%M"):
            try:
                return datetime.strptime(value.strip(), fmt).time()
            except ValueError:
                continue
    return value


def append_to_destination(records, dest_path, output_path, sheet_name=None):
    """
    Append extracted records to the end of an existing destination workbook's
    table. The destination layout (based on TestDestination.xlsm) is:

        A: (blank/Route placeholder - left empty)
        B: Date
        C: Contract
        D: (blank)
        E: Activity
        F: Material
        G: Quantity
        H: Units
        I: Source
        J: Contractor
        K: Time
        L: Location

    New rows are written starting at the first fully-empty row after the
    existing data.
    """
    wb = openpyxl.load_workbook(dest_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Find first empty row (column B used as the indicator, since col A is blank)
    next_row = ws.max_row + 1

    dest_col_map = {
        "Date": 2,        # B
        "Contract": 3,    # C
        "Activity": 5,    # E
        "Material": 6,    # F
        "Quantity": 7,    # G
        "Units": 8,       # H
        "Source": 9,      # I
        "Contractor": 10, # J
        "Time": 11,       # K
        "Location": 12,   # L
    }

    for i, rec in enumerate(records):
        row = next_row + i
        for key, col in dest_col_map.items():
            value = rec[key]
            if key == "Time":
                value = parse_time_value(value)
            ws.cell(row=row, column=col, value=value)

    wb.save(output_path)

def filter_tracker_records(records):
    """
    Keep only:
      - Material contains '90PCC'

    Exclude if Activity contains:
      - shaft
      - column
      - cap
      - bridge
      - crc
      - footing
    """

    excluded_words = [
        "shaft",
        "column",
        "cap",
        "bridge",
        "crc",
        "footing",
    ]

    filtered = []

    for rec in records:
        material = str(rec.get("Material", "")).lower()
        activity = str(rec.get("Activity", "")).lower()

        # Must contain 90PCC in Material
        if "90pcc" not in material:
            continue

        # Skip if Activity contains any excluded word
        if any(word in activity for word in excluded_words):
            continue

        filtered.append(rec)

    return filtered

def copy_row_format(ws, source_row, target_row):

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        target_cell.font = copy(source_cell.font)
        target_cell.fill = PatternFill(
            fill_type="solid",
            start_color="FFFFFF",
            end_color="FFFFFF"
        )
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)

    # ws.merge_cells(
    #     start_row=target_row,
    #     start_column=1,
    #     end_row=target_row,
    #     end_column=2
    # )

def append_to_tracker(records, tracker_path):
    print("\n=== ENTERED append_to_tracker ===")
    print(f"Tracker path = {tracker_path}")
    print(f"Records received = {len(records)}")

    wb = openpyxl.load_workbook(tracker_path)
    print("Workbook opened")

    ws = wb.active
    print(f"Sheet name = {ws.title}")

    next_row = ws.max_row + 1

    print(f"First empty row = {next_row}")

    dest_col_map = {
        "Date": 1,
        "Contract": 2,
        "Activity": 3,
        "Material": 4,
        "Quantity": 5,
        "Units": 6,
        "Source": 7,
        "Contractor": 8,
        "Time": 9,
        "Location": 10,
    }

    for i, rec in enumerate(records):
        row = next_row + i

        # Copy formatting from the previous row
        copy_row_format(ws, row - 1, row)

        print(f"Writing row {row}")

        # Copy formatting from previous row
        copy_row_format(ws, row - 1, row)

        for key, col in dest_col_map.items():
            value = rec[key]

            if key == "Time":
                value = parse_time_value(value)

            ws.cell(row=row, column=col, value=value)

    print("Saving workbook...")
    wb.save(tracker_path)
    print("Tracker saved successfully")

def send_email(file_path):
    tomorrow = datetime.now() + timedelta(days=1)

    today = tomorrow.strftime("%m/%d/%Y")

    date = (
        tomorrow.strftime("%A, %B ")
        + str(tomorrow.day)
        + tomorrow.strftime(", %Y")
    )

    recipients = [
        "rwhite@bravocoeng.com"
    ]

    msg = MIMEMultipart()
    msg["Subject"] = f"Systemwide Daily Material Placement Summary for {today}"
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = ", ".join(recipients)

    body = MIMEText(f"Good evening,\nAttached please find a systemwide daily material placement summary for {date}. The summary includes MPN's submitted by 5:30 PM.")
    msg.attach(body)

    with open(file_path, "rb") as f:
        attachment = MIMEBase("application", "vnd.ms-excel")
        attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(file_path)}"'
        )
        msg.attach(attachment)

    with smtplib.SMTP("smtp.mail.me.com", 587) as server:
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.send_message(msg)
        # server.sendmail(EMAIL_ADDRESS, recipients, msg.as_string())

def main():

    # if len(sys.argv) < 2:
    #     print("Usage:")
    #     print("python extract_mpn.py <input_file>")
    #     sys.exit(1)

    # input_path = Path(sys.argv[1])
    Tk().withdraw()

    filename = askopenfilename(
        title="Select Daily Material Placement Export",
        filetypes=[
            ("Excel files", "*.xls *.xlsx *.xlsm")
        ]
    )

    if not filename:
        return

    input_path = Path(filename)
    print(input_path)
    print(input_path.suffix)

    print("Loading:", input_path)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]

    header_row = find_header_row(ws)
    col_map = map_columns(ws, header_row)
    records = extract_records(ws, header_row, col_map)
    records = sort_records(records)

    print("\n=== APPENDING TO DESTINATION ===")
    append_to_destination(
        records,
        DEST_PATH,
        DEST_PATH
    )

    print("\n=== FILTERING TRACKER RECORDS ===")
    tracker_records = filter_tracker_records(records)

    print(f"Total records = {len(records)}")
    print(f"Tracker records = {len(tracker_records)}")

    for rec in tracker_records:
        print(rec)

    if tracker_records:
        print("\n=== CALLING append_to_tracker ===")

        append_to_tracker(
            tracker_records,
            TRACKER_PATH
        )

        print(f"Added {len(tracker_records)} rows to TestTracker2")
    else:
        print("No tracker records to append.")

    print(f"Extracted {len(records)} record(s) and appended to destination.")

    print("Sending email...")
    try:
        send_email(input_path)
        print("Email sent.")

    except Exception as e:
        print("Email failed:", e)

    for r in records:
        print(r)

    #print(f"Saved to {DEST_PATH}")
    from tkinter import messagebox

    messagebox.showinfo(
        "Finished",
        "Daily Material Placement Summary completed successfully."
    )


# if __name__ == "__main__":
#     main()
from tkinter import messagebox

if __name__ == "__main__":
    try:
        main()

    except Exception:
        messagebox.showerror(
            "Error",
            traceback.format_exc()
        )
